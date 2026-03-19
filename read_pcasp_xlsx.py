#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
读取 SPCASP 相关 Excel，将内容导出为文本。
依赖: pip install openpyxl
"""
import os
import sys

try:
    import openpyxl
except ImportError:
    print("请先安装: pip install openpyxl")
    sys.exit(1)

DIR = os.path.dirname(os.path.abspath(__file__))

def sheet_to_text(ws):
    lines = []
    for row in ws.iter_rows(values_only=True):
        line = "\t".join(str(c) if c is not None else "" for c in row)
        lines.append(line)
    return "\n".join(lines)

def main():
    want = ["SPCASP数据结构示例.xlsx", "SPCASP通信协议.xlsx"]
    all_files = os.listdir(DIR)
    for name in want:
        path = os.path.join(DIR, name)
        if not os.path.isfile(path):
            # 尝试按长度或部分名匹配
            for f in all_files:
                if not f.endswith(".xlsx"):
                    continue
                if name in f or (len(name) > 10 and name[:6] in f):
                    path = os.path.join(DIR, f)
                    name = f
                    break
        if not os.path.isfile(path):
            print(f"[跳过] 未找到: {name}")
            continue
        print(f"\n{'='*60}\n{name}\n{'='*60}")
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                print(f"\n--- Sheet: {sn} ---\n")
                print(sheet_to_text(ws))
            wb.close()
        except Exception as e:
            print(f"读取失败: {e}")
    print("\n完成.")

if __name__ == "__main__":
    main()
